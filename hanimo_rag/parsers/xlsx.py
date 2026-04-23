"""XLSX parser using openpyxl."""
from __future__ import annotations

from hanimo_rag.parsers.base import ParsedDocument, ParserBase


def _sheet_to_markdown(sheet) -> str:
    """Convert an openpyxl worksheet to a markdown table string."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    md_rows: list[str] = []
    for i, row in enumerate(rows):
        cells = [str(c) if c is not None else "" for c in row]
        md_rows.append("| " + " | ".join(cells) + " |")
        if i == 0:
            md_rows.append("| " + " | ".join(["---"] * len(cells)) + " |")

    return "\n".join(md_rows)


class XlsxParser(ParserBase):
    """Parse XLSX files, rendering each sheet as a markdown table."""

    def supported_mime_types(self) -> list[str]:
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ]

    def parse(self, file_path: str) -> ParsedDocument:
        try:
            import openpyxl
        except ImportError as e:
            return ParsedDocument(
                text="",
                metadata={"error": f"Missing dependency: {e}"},
            )

        try:
            wb = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )
        except Exception as exc:
            return ParsedDocument(text="", metadata={"error": str(exc)})

        sheet_names = wb.sheetnames
        metadata: dict = {
            "sheet_names": sheet_names,
            "sheet_count": len(sheet_names),
        }

        pages: list[str] = []
        parts: list[str] = []

        for name in sheet_names:
            ws = wb[name]
            table_md = _sheet_to_markdown(ws)
            section = f"## {name}\n\n{table_md}" if table_md else f"## {name}\n\n(empty)"
            parts.append(section)
            pages.append(section)

        wb.close()

        full_text = "\n\n".join(parts)
        return ParsedDocument(text=full_text, metadata=metadata, pages=pages)
